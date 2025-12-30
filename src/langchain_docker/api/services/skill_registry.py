"""Skill registry service implementing progressive disclosure pattern.

Based on Anthropic's Agent Skills architecture for on-demand context loading.
Reference: https://www.anthropic.com/engineering/equipping-agents-for-the-real-world-with-agent-skills

Progressive Disclosure Levels:
- Level 1 (Metadata): name, description - always in agent system prompt
- Level 2 (Core): SKILL.md body - loaded when skill is triggered
- Level 3 (Details): Additional files - loaded only as needed
"""

import logging
import re
from abc import ABC, abstractmethod
from datetime import datetime
from typing import Any, Optional

from langchain_community.utilities import SQLDatabase

from langchain_docker.api.services.demo_database import ensure_demo_database
from langchain_docker.core.config import get_database_url, is_sql_read_only

logger = logging.getLogger(__name__)


class SkillResource:
    """Additional resource file bundled with a skill."""

    def __init__(self, name: str, description: str, content: str = ""):
        self.name = name
        self.description = description
        self.content = content


class SkillScript:
    """Executable script bundled with a skill."""

    def __init__(
        self, name: str, description: str, language: str = "python", content: str = ""
    ):
        self.name = name
        self.description = description
        self.language = language
        self.content = content


class Skill(ABC):
    """Base class for skills with progressive disclosure levels.

    Progressive disclosure architecture:
    - Level 1 (Metadata): Skill id, name, description - always visible in agent prompt
    - Level 2 (Core): Main skill content - loaded on-demand via load_core()
    - Level 3 (Details): Specific resources - loaded as needed via load_details()
    """

    id: str
    name: str
    description: str
    category: str

    @abstractmethod
    def load_core(self) -> str:
        """Level 2: Load core content on-demand.

        Returns:
            Core skill content including context, guidelines, and capabilities
        """
        pass

    @abstractmethod
    def load_details(self, resource: str) -> str:
        """Level 3: Load specific detailed resources.

        Args:
            resource: Resource identifier (e.g., "samples", "examples")

        Returns:
            Detailed resource content
        """
        pass


class XLSXSkill(Skill):
    """XLSX skill for spreadsheet creation, editing, and analysis.

    Based on Anthropic's xlsx skill from https://github.com/anthropics/skills
    Provides comprehensive spreadsheet capabilities with formulas, formatting,
    data analysis, and visualization support.
    """

    def __init__(self):
        """Initialize XLSX skill."""
        self.id = "xlsx"
        self.name = "Spreadsheet Expert"
        self.description = (
            "Comprehensive spreadsheet creation, editing, and analysis with support "
            "for formulas, formatting, data analysis, and visualization"
        )
        self.category = "data"
        self.is_builtin = True

    def load_core(self) -> str:
        """Level 2: Load XLSX skill instructions.

        Returns:
            Complete skill context for spreadsheet operations
        """
        return """## XLSX Skill Activated

### Core Purpose
Create, edit, and analyze spreadsheets (.xlsx, .xlsm, .csv, .tsv) with comprehensive
support for formulas, formatting, data analysis, and visualization.

### Requirements for All Excel Files

#### Zero Formula Errors
Every Excel model MUST be delivered with ZERO formula errors:
- #REF! - Invalid cell references
- #DIV/0! - Division by zero
- #VALUE! - Wrong data type in formula
- #N/A - Value not available
- #NAME? - Unrecognized formula name

#### Preserve Existing Templates
When modifying files with established patterns:
- Study and EXACTLY match existing format, style, and conventions
- Never impose standardized formatting on files with established patterns
- Existing template conventions ALWAYS override these guidelines

### Financial Model Standards

#### Color Coding Conventions
- **Blue text (RGB: 0,0,255)**: Hardcoded inputs, numbers users will change
- **Black text (RGB: 0,0,0)**: ALL formulas and calculations
- **Green text (RGB: 0,128,0)**: Links from other worksheets in same workbook
- **Red text (RGB: 255,0,0)**: External links to other files
- **Yellow background (RGB: 255,255,0)**: Key assumptions needing attention

#### Number Formatting Rules
- **Years**: Format as text strings ("2024" not "2,024")
- **Currency**: Use $#,##0 format; specify units in headers ("Revenue ($mm)")
- **Zeros**: Format to display as "-"
- **Percentages**: Default to 0.0% format (one decimal)
- **Multiples**: Format as 0.0x for valuation multiples
- **Negative numbers**: Use parentheses (123) not minus -123

### CRITICAL: Use Formulas, Not Hardcoded Values

**Always use Excel formulas instead of calculating values in Python.**
This keeps spreadsheets dynamic and updateable.

❌ WRONG - Hardcoding:
```python
total = df['Sales'].sum()
sheet['B10'] = total  # Hardcodes value
```

✅ CORRECT - Using Formulas:
```python
sheet['B10'] = '=SUM(B2:B9)'  # Excel calculates
```

### Tool Selection

- **pandas**: Best for data analysis, bulk operations, simple data export
- **openpyxl**: Best for complex formatting, formulas, Excel-specific features

### Common Workflow

1. **Choose tool**: pandas for data, openpyxl for formulas/formatting
2. **Create/Load**: Create new workbook or load existing file
3. **Modify**: Add/edit data, formulas, and formatting
4. **Save**: Write to file
5. **Recalculate**: Run `python recalc.py output.xlsx` (MANDATORY for formulas)
6. **Verify**: Check recalc.py output for errors, fix any found

### Code Examples

#### Reading with pandas:
```python
import pandas as pd
df = pd.read_excel('file.xlsx')
all_sheets = pd.read_excel('file.xlsx', sheet_name=None)
```

#### Creating with openpyxl:
```python
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
sheet = wb.active
sheet['A1'] = 'Header'
sheet['B2'] = '=SUM(A1:A10)'  # Use formulas!
sheet['A1'].font = Font(bold=True, color='FF0000')
wb.save('output.xlsx')
```

#### Editing existing files:
```python
from openpyxl import load_workbook
wb = load_workbook('existing.xlsx')
sheet = wb.active
sheet['A1'] = 'New Value'
wb.save('modified.xlsx')
```

### Formula Verification Checklist

- [ ] Test 2-3 sample references before building full model
- [ ] Verify column mapping (column 64 = BL, not BK)
- [ ] Remember Excel rows are 1-indexed
- [ ] Check for NaN values with `pd.notna()`
- [ ] Handle division by zero in formulas
- [ ] Test formulas on 2-3 cells before applying broadly

### Best Practices

- Cell indices are 1-based in openpyxl
- Use `data_only=True` to read calculated values (WARNING: saves lose formulas)
- For large files: Use `read_only=True` or `write_only=True`
- Write minimal, concise Python code
- Add comments to cells with complex formulas
"""

    def load_details(self, resource: str) -> str:
        """Level 3: Load detailed resources.

        Args:
            resource: "recalc" for recalculation script, "examples" for code examples

        Returns:
            Detailed resource content
        """
        if resource == "recalc":
            return self._get_recalc_script()
        elif resource == "examples":
            return self._get_code_examples()
        elif resource == "formatting":
            return self._get_formatting_guide()
        else:
            return f"Unknown resource: {resource}. Available: 'recalc', 'examples', 'formatting'"

    def _get_recalc_script(self) -> str:
        """Get the recalc.py script content."""
        return '''## recalc.py - Excel Formula Recalculation Script

Recalculates all formulas in an Excel file using LibreOffice.

### Usage
```bash
python recalc.py <excel_file> [timeout_seconds]
python recalc.py output.xlsx 30
```

### Output Format
Returns JSON with error details:
```json
{
  "status": "success",           // or "errors_found"
  "total_errors": 0,             // Total error count
  "total_formulas": 42,          // Number of formulas in file
  "error_summary": {             // Only present if errors found
    "#REF!": {
      "count": 2,
      "locations": ["Sheet1!B5", "Sheet1!C10"]
    }
  }
}
```

### Script Content
```python
#!/usr/bin/env python3
"""Excel Formula Recalculation Script"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path
from openpyxl import load_workbook

def setup_libreoffice_macro():
    """Setup LibreOffice macro for recalculation if not already configured"""
    if platform.system() == 'Darwin':
        macro_dir = os.path.expanduser('~/Library/Application Support/LibreOffice/4/user/basic/Standard')
    else:
        macro_dir = os.path.expanduser('~/.config/libreoffice/4/user/basic/Standard')

    macro_file = os.path.join(macro_dir, 'Module1.xba')

    if os.path.exists(macro_file):
        with open(macro_file, 'r') as f:
            if 'RecalculateAndSave' in f.read():
                return True

    if not os.path.exists(macro_dir):
        subprocess.run(['soffice', '--headless', '--terminate_after_init'],
                      capture_output=True, timeout=10)
        os.makedirs(macro_dir, exist_ok=True)

    macro_content = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

    try:
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        return True
    except Exception:
        return False

def recalc(filename, timeout=30):
    """Recalculate formulas in Excel file and report any errors"""
    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {'error': 'Failed to setup LibreOffice macro'}

    cmd = [
        'soffice', '--headless', '--norestore',
        'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        abs_path
    ]

    result = subprocess.run(cmd, capture_output=True, text=True)

    # Check for Excel errors in the recalculated file
    try:
        wb = load_workbook(filename, data_only=True)
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()
        return {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'error_summary': {k: {'count': len(v), 'locations': v[:20]}
                            for k, v in error_details.items() if v}
        }
    except Exception as e:
        return {'error': str(e)}

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)
    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))
```
'''

    def _get_code_examples(self) -> str:
        """Get comprehensive code examples."""
        return '''## XLSX Code Examples

### 1. Create Financial Model
```python
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Financial Model"

# Headers with formatting
headers = ["", "2024", "2025", "2026"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Revenue row (blue for inputs)
ws['A2'] = "Revenue ($mm)"
ws['B2'] = 100  # Input
ws['B2'].font = Font(color='0000FF')  # Blue for input
ws['C2'] = '=B2*1.1'  # 10% growth formula
ws['D2'] = '=C2*1.1'

# Costs row
ws['A3'] = "Costs ($mm)"
ws['B3'] = '=B2*0.6'  # 60% of revenue
ws['C3'] = '=C2*0.6'
ws['D3'] = '=D2*0.6'

# Profit row
ws['A4'] = "Profit ($mm)"
ws['B4'] = '=B2-B3'
ws['C4'] = '=C2-C3'
ws['D4'] = '=D2-D3'

wb.save('financial_model.xlsx')
```

### 2. Data Analysis with pandas
```python
import pandas as pd

# Read and analyze
df = pd.read_excel('sales_data.xlsx')
print(df.describe())

# Pivot table
pivot = df.pivot_table(
    values='Amount',
    index='Region',
    columns='Quarter',
    aggfunc='sum'
)

# Save with multiple sheets
with pd.ExcelWriter('analysis.xlsx') as writer:
    df.to_excel(writer, sheet_name='Raw Data', index=False)
    pivot.to_excel(writer, sheet_name='Pivot')
```

### 3. Format Existing File
```python
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

wb = load_workbook('data.xlsx')
ws = wb.active

# Apply currency format
for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
    for cell in row:
        cell.number_format = '$#,##0'

# Highlight negative values
red_fill = PatternFill(start_color='FFCCCC', fill_type='solid')
for row in ws.iter_rows(min_row=2):
    for cell in row:
        if isinstance(cell.value, (int, float)) and cell.value < 0:
            cell.fill = red_fill

wb.save('formatted_data.xlsx')
```

### 4. Conditional Formulas
```python
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# Sample data
ws['A1'] = 'Score'
ws['B1'] = 'Grade'
for i, score in enumerate([95, 82, 67, 55, 88], 2):
    ws[f'A{i}'] = score
    # Nested IF for grading
    ws[f'B{i}'] = f'=IF(A{i}>=90,"A",IF(A{i}>=80,"B",IF(A{i}>=70,"C",IF(A{i}>=60,"D","F"))))'

wb.save('grades.xlsx')
```
'''

    def _get_formatting_guide(self) -> str:
        """Get detailed formatting guide."""
        return '''## XLSX Formatting Guide

### Color Codes (RGB)
| Purpose | Color | RGB Code |
|---------|-------|----------|
| Hardcoded inputs | Blue | 0000FF |
| Formulas | Black | 000000 |
| Internal links | Green | 008000 |
| External links | Red | FF0000 |
| Key assumptions | Yellow bg | FFFF00 |

### Number Formats
| Type | Format Code | Example |
|------|-------------|---------|
| Currency | $#,##0 | $1,234 |
| Currency (neg) | $#,##0;($#,##0) | ($1,234) |
| Percentage | 0.0% | 12.5% |
| Multiple | 0.0x | 2.5x |
| Date | YYYY-MM-DD | 2024-12-29 |
| Text Year | @ | 2024 |

### openpyxl Formatting Code
```python
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Font styling
cell.font = Font(
    name='Calibri',
    size=11,
    bold=True,
    italic=False,
    color='0000FF'  # Blue
)

# Fill (background)
cell.fill = PatternFill(
    start_color='FFFF00',  # Yellow
    fill_type='solid'
)

# Alignment
cell.alignment = Alignment(
    horizontal='center',  # left, center, right
    vertical='center',    # top, center, bottom
    wrap_text=True
)

# Borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
cell.border = thin_border

# Number format
cell.number_format = '$#,##0'
```

### Column Width & Row Height
```python
# Set column width
ws.column_dimensions['A'].width = 20

# Set row height
ws.row_dimensions[1].height = 30

# Auto-fit (approximate)
for column in ws.columns:
    max_length = max(len(str(cell.value or '')) for cell in column)
    ws.column_dimensions[column[0].column_letter].width = max_length + 2
```
'''


class SQLSkill(Skill):
    """SQL skill with database schema progressive disclosure.

    Provides database querying capabilities with on-demand schema loading.
    This keeps the base agent context lightweight while allowing full
    database introspection when needed.
    """

    def __init__(self, db_url: Optional[str] = None, read_only: Optional[bool] = None):
        """Initialize SQL skill.

        Args:
            db_url: Database URL (defaults to DATABASE_URL env var)
            read_only: Whether to enforce read-only mode (defaults to SQL_READ_ONLY env var)
        """
        self.id = "write_sql"
        self.name = "SQL Query Expert"
        self.description = "Write and execute SQL queries against the database"
        self.category = "database"
        self.db_url = db_url or get_database_url()
        self.read_only = read_only if read_only is not None else is_sql_read_only()
        self._db: Optional[SQLDatabase] = None

    def _get_db(self) -> SQLDatabase:
        """Get or create SQLDatabase instance.

        Returns:
            SQLDatabase wrapper for the configured database
        """
        if self._db is None:
            # Ensure demo database exists for SQLite
            if self.db_url.startswith("sqlite:///"):
                ensure_demo_database(self.db_url)
            self._db = SQLDatabase.from_uri(self.db_url)
        return self._db

    def load_core(self) -> str:
        """Level 2: Load database schema and SQL guidelines.

        Returns:
            Complete skill context including tables, schema, and guidelines
        """
        db = self._get_db()
        tables = db.get_usable_table_names()
        schema = db.get_table_info()
        dialect = db.dialect

        read_only_note = ""
        if self.read_only:
            read_only_note = """
### Read-Only Mode
This database is in READ-ONLY mode. Only SELECT queries are allowed.
INSERT, UPDATE, DELETE, and other write operations will be rejected.
"""

        return f"""## SQL Skill Activated

### Database Information
- Dialect: {dialect}
- Available Tables: {', '.join(tables)}

### Database Schema
{schema}
{read_only_note}
### Guidelines
- Always use explicit column names (avoid SELECT *)
- Use appropriate JOINs for related tables
- Use LIMIT to prevent large result sets
- Format queries for readability
- Explain query results clearly to users

### Query Patterns
- Count records: SELECT COUNT(*) FROM table_name
- Filter data: SELECT ... WHERE condition
- Join tables: SELECT ... FROM t1 JOIN t2 ON t1.id = t2.foreign_id
- Aggregate: SELECT column, SUM(value) FROM ... GROUP BY column
- Sort: SELECT ... ORDER BY column DESC LIMIT n
"""

    def load_details(self, resource: str) -> str:
        """Level 3: Load sample rows or query examples.

        Args:
            resource: "samples" for sample rows, "examples" for query examples

        Returns:
            Detailed resource content
        """
        db = self._get_db()

        if resource == "samples":
            tables = db.get_usable_table_names()
            samples = []
            for table in tables[:5]:  # Limit to first 5 tables
                try:
                    result = db.run(f"SELECT * FROM {table} LIMIT 3")
                    samples.append(f"### {table}\n{result}")
                except Exception as e:
                    samples.append(f"### {table}\nError fetching samples: {e}")
            return "\n\n".join(samples)

        elif resource == "examples":
            return """### Example Queries

1. **List all customers from a city**
```sql
SELECT name, email FROM customers WHERE city = 'New York'
```

2. **Get top customers by order total**
```sql
SELECT c.name, SUM(o.total) as total_spent
FROM customers c
JOIN orders o ON c.id = o.customer_id
GROUP BY c.id
ORDER BY total_spent DESC
LIMIT 5
```

3. **Find products in a category**
```sql
SELECT name, price FROM products
WHERE category = 'Electronics'
ORDER BY price DESC
```

4. **Recent orders with customer info**
```sql
SELECT c.name, p.name as product, o.quantity, o.total, o.order_date
FROM orders o
JOIN customers c ON o.customer_id = c.id
JOIN products p ON o.product_id = p.id
ORDER BY o.order_date DESC
LIMIT 10
```
"""

        else:
            return f"Unknown resource: {resource}. Available: 'samples', 'examples'"

    def execute_query(self, query: str) -> str:
        """Execute a SQL query with read-only enforcement.

        Args:
            query: SQL query string

        Returns:
            Query results or error message
        """
        db = self._get_db()

        # Enforce read-only mode
        if self.read_only:
            query_upper = query.strip().upper()
            write_keywords = ["INSERT", "UPDATE", "DELETE", "DROP", "CREATE", "ALTER", "TRUNCATE"]
            for keyword in write_keywords:
                if query_upper.startswith(keyword):
                    return f"Error: {keyword} operations are not allowed in read-only mode. Only SELECT queries are permitted."

        try:
            result = db.run(query)
            return result
        except Exception as e:
            return f"Query error: {str(e)}"

    def list_tables(self) -> str:
        """List all available tables.

        Returns:
            Comma-separated list of table names
        """
        db = self._get_db()
        tables = db.get_usable_table_names()
        return ", ".join(tables)

    def get_table_schema(self, table_name: str) -> str:
        """Get schema for a specific table.

        Args:
            table_name: Name of the table

        Returns:
            Table schema information
        """
        db = self._get_db()
        try:
            return db.get_table_info([table_name])
        except Exception as e:
            return f"Error getting schema for {table_name}: {str(e)}"


class CustomSkill(Skill):
    """User-created skill following SKILL.md format.

    Supports the full progressive disclosure pattern with:
    - YAML frontmatter for metadata
    - Markdown body for core instructions
    - Additional resource files
    - Bundled executable scripts
    """

    def __init__(
        self,
        skill_id: str,
        name: str,
        description: str,
        category: str = "general",
        version: str = "1.0.0",
        author: Optional[str] = None,
        core_content: str = "",
        resources: Optional[list[SkillResource]] = None,
        scripts: Optional[list[SkillScript]] = None,
    ):
        """Initialize a custom skill.

        Args:
            skill_id: Unique skill identifier
            name: Human-readable skill name
            description: Brief description of what the skill does
            category: Skill category
            version: Skill version
            author: Skill author
            core_content: Main skill instructions (Level 2)
            resources: Additional resource files (Level 3)
            scripts: Executable scripts
        """
        self.id = skill_id
        self.name = name
        self.description = description
        self.category = category
        self.version = version
        self.author = author
        self._core_content = core_content
        self._resources = resources or []
        self._scripts = scripts or []
        self.is_builtin = False
        self.created_at = datetime.utcnow().isoformat()
        self.updated_at = self.created_at

    def load_core(self) -> str:
        """Level 2: Load core skill content.

        Returns:
            Main skill instructions in markdown format
        """
        # Build the skill context
        content = f"""## {self.name} Skill Activated

{self._core_content}
"""
        # Add references to available resources
        if self._resources:
            content += "\n### Additional Resources\n"
            content += "The following resources are available for more details:\n"
            for resource in self._resources:
                content += f"- `{resource.name}`: {resource.description}\n"

        # Add references to available scripts
        if self._scripts:
            content += "\n### Available Scripts\n"
            content += "The following scripts can be executed:\n"
            for script in self._scripts:
                content += f"- `{script.name}` ({script.language}): {script.description}\n"

        return content

    def load_details(self, resource: str) -> str:
        """Level 3: Load a specific resource.

        Args:
            resource: Resource name to load

        Returns:
            Resource content or error message
        """
        # Check if it's a resource file
        for res in self._resources:
            if res.name == resource:
                return f"## {res.name}\n\n{res.content}"

        # Check if it's a script
        for script in self._scripts:
            if script.name == resource:
                return f"## {script.name} ({script.language})\n\n```{script.language}\n{script.content}\n```"

        available = [r.name for r in self._resources] + [s.name for s in self._scripts]
        return f"Unknown resource: {resource}. Available: {', '.join(available)}"

    def execute_script(self, script_name: str, args: dict[str, Any] = None) -> str:
        """Execute a bundled script.

        Args:
            script_name: Name of the script to execute
            args: Arguments to pass to the script

        Returns:
            Script output or error message
        """
        script = None
        for s in self._scripts:
            if s.name == script_name:
                script = s
                break

        if not script:
            available = [s.name for s in self._scripts]
            return f"Unknown script: {script_name}. Available: {', '.join(available)}"

        # For now, return the script content
        # In production, this would execute in a sandbox
        return f"Script execution not yet implemented. Script content:\n{script.content}"

    def update(
        self,
        name: Optional[str] = None,
        description: Optional[str] = None,
        category: Optional[str] = None,
        version: Optional[str] = None,
        author: Optional[str] = None,
        core_content: Optional[str] = None,
        resources: Optional[list[SkillResource]] = None,
        scripts: Optional[list[SkillScript]] = None,
    ) -> None:
        """Update skill properties.

        Args:
            name: New name (optional)
            description: New description (optional)
            category: New category (optional)
            version: New version (optional)
            author: New author (optional)
            core_content: New core content (optional)
            resources: New resources (optional)
            scripts: New scripts (optional)
        """
        if name is not None:
            self.name = name
        if description is not None:
            self.description = description
        if category is not None:
            self.category = category
        if version is not None:
            self.version = version
        if author is not None:
            self.author = author
        if core_content is not None:
            self._core_content = core_content
        if resources is not None:
            self._resources = resources
        if scripts is not None:
            self._scripts = scripts
        self.updated_at = datetime.utcnow().isoformat()

    def to_dict(self) -> dict[str, Any]:
        """Convert skill to dictionary representation.

        Returns:
            Dictionary with all skill data
        """
        return {
            "id": self.id,
            "name": self.name,
            "description": self.description,
            "category": self.category,
            "version": self.version,
            "author": self.author,
            "is_builtin": self.is_builtin,
            "core_content": self._core_content,
            "resources": [
                {"name": r.name, "description": r.description, "content": r.content}
                for r in self._resources
            ],
            "scripts": [
                {
                    "name": s.name,
                    "description": s.description,
                    "language": s.language,
                    "content": s.content,
                }
                for s in self._scripts
            ],
            "created_at": self.created_at,
            "updated_at": self.updated_at,
        }

    @classmethod
    def from_skill_md(cls, content: str) -> "CustomSkill":
        """Parse a SKILL.md file content into a CustomSkill.

        Args:
            content: Full SKILL.md content with YAML frontmatter

        Returns:
            CustomSkill instance
        """
        # Parse YAML frontmatter
        frontmatter_match = re.match(r"^---\s*\n(.*?)\n---\s*\n(.*)$", content, re.DOTALL)

        if not frontmatter_match:
            raise ValueError("Invalid SKILL.md format: missing YAML frontmatter")

        import yaml

        frontmatter = yaml.safe_load(frontmatter_match.group(1))
        body = frontmatter_match.group(2).strip()

        # Generate ID from name if not provided
        skill_id = frontmatter.get("id") or re.sub(
            r"[^a-z0-9]+", "_", frontmatter["name"].lower()
        ).strip("_")

        return cls(
            skill_id=skill_id,
            name=frontmatter["name"],
            description=frontmatter["description"],
            category=frontmatter.get("category", "general"),
            version=frontmatter.get("version", "1.0.0"),
            author=frontmatter.get("author"),
            core_content=body,
        )

    def to_skill_md(self) -> str:
        """Export skill as SKILL.md format.

        Returns:
            SKILL.md content with YAML frontmatter
        """
        import yaml

        frontmatter = {
            "name": self.name,
            "description": self.description,
            "category": self.category,
            "version": self.version,
        }
        if self.author:
            frontmatter["author"] = self.author

        return f"---\n{yaml.dump(frontmatter, default_flow_style=False)}---\n\n{self._core_content}"


class SkillRegistry:
    """Registry of available skills for progressive disclosure.

    Manages skill metadata (Level 1) and provides access to
    on-demand skill loading (Levels 2 and 3).

    Supports both built-in skills (SQLSkill, etc.) and custom
    user-created skills following the SKILL.md format.
    """

    def __init__(self):
        """Initialize skill registry with built-in skills."""
        self._skills: dict[str, Skill] = {}
        self._custom_skills: dict[str, CustomSkill] = {}
        self._register_builtin_skills()

    def _register_builtin_skills(self) -> None:
        """Register all built-in skills."""
        # SQL skill
        sql_skill = SQLSkill()
        sql_skill.is_builtin = True  # Mark as built-in
        self.register(sql_skill)

        # XLSX skill (from Anthropic skills repository)
        xlsx_skill = XLSXSkill()
        self.register(xlsx_skill)

    def register(self, skill: Skill) -> None:
        """Register a skill.

        Args:
            skill: Skill instance to register
        """
        self._skills[skill.id] = skill
        logger.info(f"Registered skill: {skill.id} ({skill.name})")

    def get_skill(self, skill_id: str) -> Optional[Skill]:
        """Get a skill by ID.

        Args:
            skill_id: Skill identifier

        Returns:
            Skill if found, None otherwise
        """
        return self._skills.get(skill_id)

    def list_skills(self) -> list[dict[str, Any]]:
        """List all available skills (Level 1 metadata).

        Returns:
            List of skill metadata dictionaries
        """
        return [
            {
                "id": skill.id,
                "name": skill.name,
                "description": skill.description,
                "category": skill.category,
            }
            for skill in self._skills.values()
        ]

    def get_skill_summary(self) -> str:
        """Get a summary of available skills for agent prompts.

        Returns:
            Formatted string listing available skills
        """
        lines = ["Available skills (use load_sql_skill to activate):"]
        for skill in self._skills.values():
            lines.append(f"- {skill.id}: {skill.description}")
        return "\n".join(lines)

    def load_skill(self, skill_id: str) -> str:
        """Load a skill's core content (Level 2).

        Args:
            skill_id: Skill identifier

        Returns:
            Skill core content or error message
        """
        skill = self.get_skill(skill_id)
        if not skill:
            available = ", ".join(self._skills.keys())
            return f"Unknown skill: {skill_id}. Available skills: {available}"

        return skill.load_core()

    def load_skill_details(self, skill_id: str, resource: str) -> str:
        """Load a skill's detailed resources (Level 3).

        Args:
            skill_id: Skill identifier
            resource: Resource identifier

        Returns:
            Detailed resource content or error message
        """
        skill = self.get_skill(skill_id)
        if not skill:
            available = ", ".join(self._skills.keys())
            return f"Unknown skill: {skill_id}. Available skills: {available}"

        return skill.load_details(resource)

    # Custom Skill CRUD Operations

    def create_custom_skill(
        self,
        name: str,
        description: str,
        core_content: str,
        skill_id: Optional[str] = None,
        category: str = "general",
        version: str = "1.0.0",
        author: Optional[str] = None,
        resources: Optional[list[dict]] = None,
        scripts: Optional[list[dict]] = None,
    ) -> CustomSkill:
        """Create a new custom skill.

        Args:
            name: Skill name
            description: Skill description
            core_content: Main skill instructions
            skill_id: Custom ID (auto-generated if not provided)
            category: Skill category
            version: Skill version
            author: Skill author
            resources: Additional resource files
            scripts: Executable scripts

        Returns:
            Created CustomSkill instance

        Raises:
            ValueError: If skill_id already exists
        """
        # Generate ID from name if not provided
        if not skill_id:
            skill_id = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")

        # Check for duplicates
        if skill_id in self._skills:
            raise ValueError(f"Skill with ID '{skill_id}' already exists")

        # Convert resource/script dicts to objects
        skill_resources = []
        if resources:
            for r in resources:
                skill_resources.append(
                    SkillResource(
                        name=r["name"],
                        description=r.get("description", ""),
                        content=r.get("content", ""),
                    )
                )

        skill_scripts = []
        if scripts:
            for s in scripts:
                skill_scripts.append(
                    SkillScript(
                        name=s["name"],
                        description=s.get("description", ""),
                        language=s.get("language", "python"),
                        content=s.get("content", ""),
                    )
                )

        # Create the skill
        skill = CustomSkill(
            skill_id=skill_id,
            name=name,
            description=description,
            category=category,
            version=version,
            author=author,
            core_content=core_content,
            resources=skill_resources,
            scripts=skill_scripts,
        )

        # Register it
        self._skills[skill_id] = skill
        self._custom_skills[skill_id] = skill
        logger.info(f"Created custom skill: {skill_id} ({name})")

        return skill

    def update_custom_skill(
        self,
        skill_id: str,
        name: Optional[str] = None,
        description: Optional[str] = None,
        core_content: Optional[str] = None,
        category: Optional[str] = None,
        version: Optional[str] = None,
        author: Optional[str] = None,
        resources: Optional[list[dict]] = None,
        scripts: Optional[list[dict]] = None,
    ) -> CustomSkill:
        """Update an existing custom skill.

        Args:
            skill_id: Skill ID to update
            name: New name (optional)
            description: New description (optional)
            core_content: New core content (optional)
            category: New category (optional)
            version: New version (optional)
            author: New author (optional)
            resources: New resources (optional)
            scripts: New scripts (optional)

        Returns:
            Updated CustomSkill instance

        Raises:
            ValueError: If skill not found or is a built-in skill
        """
        skill = self._custom_skills.get(skill_id)
        if not skill:
            if skill_id in self._skills:
                raise ValueError(f"Cannot update built-in skill: {skill_id}")
            raise ValueError(f"Skill not found: {skill_id}")

        # Convert resource/script dicts to objects if provided
        skill_resources = None
        if resources is not None:
            skill_resources = []
            for r in resources:
                skill_resources.append(
                    SkillResource(
                        name=r["name"],
                        description=r.get("description", ""),
                        content=r.get("content", ""),
                    )
                )

        skill_scripts = None
        if scripts is not None:
            skill_scripts = []
            for s in scripts:
                skill_scripts.append(
                    SkillScript(
                        name=s["name"],
                        description=s.get("description", ""),
                        language=s.get("language", "python"),
                        content=s.get("content", ""),
                    )
                )

        # Update the skill
        skill.update(
            name=name,
            description=description,
            category=category,
            version=version,
            author=author,
            core_content=core_content,
            resources=skill_resources,
            scripts=skill_scripts,
        )

        logger.info(f"Updated custom skill: {skill_id}")
        return skill

    def delete_custom_skill(self, skill_id: str) -> bool:
        """Delete a custom skill.

        Args:
            skill_id: Skill ID to delete

        Returns:
            True if deleted

        Raises:
            ValueError: If skill not found or is a built-in skill
        """
        if skill_id not in self._custom_skills:
            if skill_id in self._skills:
                raise ValueError(f"Cannot delete built-in skill: {skill_id}")
            raise ValueError(f"Skill not found: {skill_id}")

        del self._skills[skill_id]
        del self._custom_skills[skill_id]
        logger.info(f"Deleted custom skill: {skill_id}")
        return True

    def get_skill_full(self, skill_id: str) -> Optional[dict[str, Any]]:
        """Get full skill information including content.

        Args:
            skill_id: Skill identifier

        Returns:
            Full skill data or None if not found
        """
        skill = self.get_skill(skill_id)
        if not skill:
            return None

        if isinstance(skill, CustomSkill):
            return skill.to_dict()

        # For built-in skills, return basic info
        return {
            "id": skill.id,
            "name": skill.name,
            "description": skill.description,
            "category": skill.category,
            "is_builtin": True,
            "core_content": skill.load_core(),
            "resources": [],
            "scripts": [],
        }

    def list_skills_full(self) -> list[dict[str, Any]]:
        """List all skills with full metadata.

        Returns:
            List of skill dictionaries with metadata
        """
        skills = []
        for skill in self._skills.values():
            is_builtin = not isinstance(skill, CustomSkill)
            skill_data = {
                "id": skill.id,
                "name": skill.name,
                "description": skill.description,
                "category": skill.category,
                "is_builtin": is_builtin,
            }

            if isinstance(skill, CustomSkill):
                skill_data["version"] = skill.version
                skill_data["author"] = skill.author
                skill_data["created_at"] = skill.created_at
                skill_data["updated_at"] = skill.updated_at

            skills.append(skill_data)

        return skills

    def import_skill_md(self, content: str) -> CustomSkill:
        """Import a skill from SKILL.md format.

        Args:
            content: Full SKILL.md content with YAML frontmatter

        Returns:
            Created CustomSkill instance
        """
        skill = CustomSkill.from_skill_md(content)

        # Check for duplicates
        if skill.id in self._skills:
            raise ValueError(f"Skill with ID '{skill.id}' already exists")

        self._skills[skill.id] = skill
        self._custom_skills[skill.id] = skill
        logger.info(f"Imported skill from SKILL.md: {skill.id}")

        return skill

    def export_skill_md(self, skill_id: str) -> str:
        """Export a skill as SKILL.md format.

        Args:
            skill_id: Skill ID to export

        Returns:
            SKILL.md content

        Raises:
            ValueError: If skill not found or is a built-in skill
        """
        skill = self._custom_skills.get(skill_id)
        if not skill:
            if skill_id in self._skills:
                raise ValueError(f"Cannot export built-in skill: {skill_id}")
            raise ValueError(f"Skill not found: {skill_id}")

        return skill.to_skill_md()
